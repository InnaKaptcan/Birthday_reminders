import pandas as pd
from datetime import datetime as dt
import openpyxl
import os

from Lead_preferences import departments
from Days_of_interest import todays_file_name, days

pd.options.mode.chained_assignment = None

# ПОДГОТОВИТЕЛЬНЫЕ РАБОТЫ-------------------> таблица birthdays7dep

# Файл лежит в папке Downloads
os.chdir("/Users/innakaptcan/Downloads")

# Загрузка таблички с ДР в датафрейм. Если в загрузках еще нет таблички в формате xlsx, то сначала вызов программы
# конвертации.
try:
    file_from1C = pd.read_excel(todays_file_name + '.xlsx', dtype="object")
except FileNotFoundError:
    with open("/Users/innakaptcan/PycharmProjects/Birthday_reminders/Converting_into_xlsx.py") as f:
        exec(f.read())
    file_from1C = pd.read_excel(todays_file_name + '.xlsx', dtype="object")

file_from1C = pd.read_excel(todays_file_name + '.xlsx', dtype="object")
# Чтобы восстановить нули в начале кода подразделения, загрузим табличку с ДР в openpyxl workbook
openpyxl_wb = openpyxl.load_workbook(todays_file_name + '.xlsx')

# Пройдемся по строчкам openpyxl_wb и найдем ячейки с нестанртным форматированием в колонке "Код".
# Добавим нули в выявленных ячейках в датафрейме.
# Чтобы определить, кого оповещать о ДР руководителя, обрежим последнюю цифру кода подразделения у людей, у которых
# в колонке 'Руковод' звездочка.
# В openpyxl_wb номер колонки на 1 больше, а номер строки на 2 больше
kod_column_number = file_from1C.columns.get_loc('Код')
lead_column_number = file_from1C.columns.get_loc('Руковод.')
for i in range(3, len(file_from1C) + 3):
    file_from1C.iloc[i-3, kod_column_number] = str(file_from1C.iloc[i-3, kod_column_number]).replace(' ', '')
    if openpyxl_wb['TDSheet'].cell(column=kod_column_number + 1, row=i-1).number_format in ['00\"   \"', '000\"  \"']:
        file_from1C.iloc[i-3, kod_column_number] = '0' + file_from1C.iloc[i-3, kod_column_number]
    if openpyxl_wb['TDSheet'].cell(column=lead_column_number + 1, row=i-1).value == '*':
        file_from1C.iloc[i-3, kod_column_number] = file_from1C.iloc[i-3, kod_column_number][:-1]

# Удаление декоративных срочек.
file_from1C = file_from1C.dropna(axis=0, how='all')


file_from1C = file_from1C.loc[file_from1C['Код'].isin(departments)]

# Приводим колонки в нужный формат: обрезаем год у даты рождения и заменяем ее на текущий год, возраст в целочисленный формат
file_from1C['День'] = file_from1C['День'].astype('string').map(lambda x: dt.strptime(x[:6] + '23', '%d.%m.%y').date())
file_from1C['Возраст'] = file_from1C['Возраст'].astype('int')


birthdays7dep = file_from1C.loc[file_from1C['День'].isin(days)]
birthdays7dep['Фамилия имя отчество сотрудника'] = birthdays7dep['Фамилия имя отчество сотрудника'].map(
    lambda x: x.split(' ')[0] + ' ' + x.split(' ')[1])
